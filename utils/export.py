import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import datetime
import os
from io import BytesIO

def create_excel_workbook():
    """Создаёт новый workbook с базовыми настройками"""
    wb = openpyxl.Workbook()
    
    # Настройки по умолчанию
    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
    
    return wb

def style_header(cell, text):
    """Применяет стиль к заголовку"""
    cell.value = text
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def style_cell(cell):
    """Применяет стиль к ячейке"""
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def auto_adjust_columns(worksheet):
    """Автоматически подгоняет ширину колонок"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def export_orders_to_excel(orders, filename=None):
    """
    Экспортирует заказы в Excel
    orders: список заказов из базы данных
    filename: если None, возвращает BytesIO для отправки в Telegram
    """
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Заказы"
    
    # Заголовки
    headers = [
        "ID заказа", "Дата", "Время", "Клиент", "Телефон", 
        "Адрес", "Подъезд", "Этаж", "Квартира", "Домофон",
        "Мешков", "Цена", "Статус", "Курьер ID", "Создан"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        style_header(cell, header)
    
    # Данные
    for row, order in enumerate(orders, 2):
        # order: (id, user_id, name, phone, street, entrance, floor, apt, intercom, date, time, bags, price, status, courier_id, created)
        
        # Формируем полный адрес
        full_address = order[4]  # street
        data = [
            order[0],  # ID
            order[9],  # date
            order[10],  # time
            order[2],  # name
            order[3],  # phone
            full_address,
            order[5] or '',  # entrance
            order[6] or '',  # floor
            order[7] or '',  # apartment
            order[8] or '',  # intercom
            order[11],  # bags
            order[12],  # price
            order[13],  # status
            order[14] or '',  # courier_id
            order[15].strftime("%d.%m.%Y %H:%M") if order[15] else ''  # created
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            style_cell(cell)
    
    auto_adjust_columns(ws)
    
    # Статистика внизу
    total_row = len(orders) + 3
    ws.cell(row=total_row, column=1, value=f"Всего заказов: {len(orders)}")
    ws.cell(row=total_row, column=2, value=f"Общая сумма: {sum(o[12] for o in orders)} ₽")
    ws.cell(row=total_row, column=3, value=f"Всего мешков: {sum(o[11] for o in orders)}")
    
    return save_workbook(wb, filename, "orders.xlsx")

def export_clients_to_excel(clients, filename=None):
    """Экспортирует клиентов в Excel"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Клиенты"
    
    headers = [
        "ID", "Username", "Имя", "Фамилия", "Телефон",
        "Адрес", "Подъезд", "Этаж", "Квартира", "Домофон",
        "Дата регистрации"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        style_header(cell, header)
    
    for row, client in enumerate(clients, 2):
        # client: (user_id, username, first_name, last_name, phone, street, entrance, floor, apt, intercom, reg_date)
        data = [
            client[0],  # user_id
            f"@{client[1]}" if client[1] else "",
            client[2] or "",
            client[3] or "",
            client[4] or "",
            client[5] or "",
            client[6] or "",
            client[7] or "",
            client[8] or "",
            client[9] or "",
            client[10].strftime("%d.%m.%Y %H:%M") if client[10] else ""
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            style_cell(cell)
    
    auto_adjust_columns(ws)
    return save_workbook(wb, filename, "clients.xlsx")

def export_stats_to_excel(stats_data, filename=None):
    """Экспортирует статистику в Excel"""
    wb = create_excel_workbook()
    
    # Лист с общей статистикой
    ws1 = wb.active
    ws1.title = "Общая статистика"
    
    stats = [
        ["Показатель", "Значение"],
        ["Всего заказов", stats_data.get('total_orders', 0)],
        ["Новых заказов", stats_data.get('new_orders', 0)],
        ["Подтверждённых", stats_data.get('confirmed_orders', 0)],
        ["Выполненных", stats_data.get('completed_orders', 0)],
        ["Отменённых", stats_data.get('cancelled_orders', 0)],
        ["Всего клиентов", stats_data.get('total_clients', 0)],
        ["Активных клиентов", stats_data.get('active_clients', 0)],
        ["Общая выручка", f"{stats_data.get('total_revenue', 0)} ₽"],
        ["Средний чек", f"{stats_data.get('avg_check', 0)} ₽"],
        ["Всего мешков", stats_data.get('total_bags', 0)]
    ]
    
    for row, (param, value) in enumerate(stats, 1):
        ws1.cell(row=row, column=1, value=param)
        ws1.cell(row=row, column=2, value=value)
    
    # Лист с ежедневной статистикой
    if 'daily_stats' in stats_data:
        ws2 = wb.create_sheet("По дням")
        headers = ["Дата", "Заказов", "Выручка", "Мешков"]
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col)
            style_header(cell, header)
        
        for row, day in enumerate(stats_data['daily_stats'], 2):
            ws2.cell(row=row, column=1, value=day['date'])
            ws2.cell(row=row, column=2, value=day['orders'])
            ws2.cell(row=row, column=3, value=f"{day['revenue']} ₽")
            ws2.cell(row=row, column=4, value=day['bags'])
    
    return save_workbook(wb, filename, "stats.xlsx")

def export_blacklist_to_excel(blacklist, filename=None):
    """Экспортирует чёрный список в Excel"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Чёрный список"
    
    headers = ["ID", "Имя", "Username", "Телефон", "Причина", "Дата добавления"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        style_header(cell, header)
    
    for row, user in enumerate(blacklist, 2):
        # user: (user_id, reason, added_date, username, first_name, phone)
        data = [
            user[0],  # user_id
            user[4] or "",
            f"@{user[3]}" if user[3] else "",
            user[5] or "",
            user[1] or "",
            user[2].strftime("%d.%m.%Y %H:%M") if user[2] else ""
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            style_cell(cell)
    
    auto_adjust_columns(ws)
    return save_workbook(wb, filename, "blacklist.xlsx")

def export_messages_to_excel(messages, filename=None):
    """Экспортирует сообщения из мини-мессенджера"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Сообщения"
    
    headers = ["ID", "От кого", "Текст", "Ответ админа", "Статус", "Дата"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        style_header(cell, header)
    
    for row, msg in enumerate(messages, 2):
        # msg: (id, user_id, username, first_name, phone, user_msg, reply, status, created)
        data = [
            msg[0],  # id
            f"{msg[3] or ''} (@{msg[2] or ''})" if msg[2] else msg[3] or f"ID {msg[1]}",
            msg[5] or "",
            msg[6] or "",
            msg[7] or "",
            msg[8].strftime("%d.%m.%Y %H:%M") if msg[8] else ""
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            style_cell(cell)
    
    auto_adjust_columns(ws)
    return save_workbook(wb, filename, "messages.xlsx")

def save_workbook(wb, filename=None, default_name="export.xlsx"):
    """Сохраняет workbook в файл или возвращает BytesIO"""
    if filename:
        wb.save(filename)
        return filename
    else:
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output